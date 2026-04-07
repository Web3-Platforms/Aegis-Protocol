#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime guidance only
    raise SystemExit(
        "openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_MD = REPO_ROOT / "docs/project-management/AEGIS_PROJECT_WORKBOOK.md"
OPERATING_MODEL_MD = REPO_ROOT / "docs/project-management/OPERATING_MODEL.md"
OUTPUT_XLSX = REPO_ROOT / "docs/project-management/PROJECT_MANAGEMENT.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def read_lines(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8").splitlines()


def split_table_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def find_heading(lines: list[str], heading: str) -> int:
    for idx, line in enumerate(lines):
        if line.strip() == heading:
            return idx
    raise ValueError(f"Heading not found: {heading}")


def find_heading_prefix(lines: list[str], prefix: str) -> int | None:
    for idx, line in enumerate(lines):
        if line.strip().startswith(prefix):
            return idx
    return None


def parse_table(lines: list[str], start_idx: int) -> tuple[list[str], list[list[str]], int]:
    headers = split_table_row(lines[start_idx])
    idx = start_idx + 2  # skip separator line
    rows: list[list[str]] = []

    while idx < len(lines) and lines[idx].lstrip().startswith("|"):
        row = split_table_row(lines[idx])
        if any(cell for cell in row):
            rows.append(row)
        idx += 1

    return headers, rows, idx


def parse_table_after_heading(lines: list[str], heading: str) -> tuple[list[str], list[list[str]]]:
    idx = find_heading(lines, heading) + 1
    while idx < len(lines) and not lines[idx].lstrip().startswith("|"):
        idx += 1
    headers, rows, _ = parse_table(lines, idx)
    return headers, rows


def parse_optional_table_after_heading_prefix(
    lines: list[str], prefix: str
) -> tuple[list[str] | None, list[list[str]]]:
    idx = find_heading_prefix(lines, prefix)
    if idx is None:
        return None, []

    idx += 1
    while idx < len(lines) and not lines[idx].lstrip().startswith("|"):
        idx += 1

    if idx >= len(lines):
        return None, []

    headers, rows, _ = parse_table(lines, idx)
    return headers, rows


def parse_backlog(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    epic_re = re.compile(r"^### (AEGIS-E\d+) - (.+)$")
    backlog_rows: list[list[str]] = []
    idx = 0

    while idx < len(lines):
        match = epic_re.match(lines[idx].strip())
        if not match:
            idx += 1
            continue

        epic_key, epic_name = match.groups()
        goal = ""
        scan = idx + 1

        while scan < len(lines):
            text = lines[scan].strip()
            if text.startswith("**Goal:**"):
                goal = text.replace("**Goal:**", "", 1).strip()
                break
            if text.startswith("### ") or text.startswith("## "):
                break
            scan += 1

        while scan < len(lines) and not lines[scan].lstrip().startswith("| Key | Type |"):
            scan += 1

        if scan >= len(lines):
            idx += 1
            continue

        headers, rows, end_idx = parse_table(lines, scan)
        for row in rows:
            backlog_rows.append([epic_key, epic_name, goal] + row)
        idx = end_idx

    backlog_headers = ["Epic Key", "Epic", "Goal"] + headers
    return backlog_headers, backlog_rows


def parse_numbered_list_after_heading(lines: list[str], heading: str) -> tuple[list[str], list[list[str]]]:
    start_idx = find_heading(lines, heading) + 1
    item_re = re.compile(r"^(\d+)\.\s+(.*)$")
    rows: list[list[str]] = []

    idx = start_idx
    while idx < len(lines):
        text = lines[idx].strip()
        if text.startswith("## "):
            break
        match = item_re.match(text)
        if match:
            rows.append([match.group(1), match.group(2)])
        idx += 1

    return ["Order", "Item"], rows


def parse_weekly_update(lines: list[str]) -> list[list[str]]:
    start_idx = find_heading(lines, "## Weekly update template") + 1
    rows: list[list[str]] = []
    current_section: str | None = None

    idx = start_idx
    while idx < len(lines):
        text = lines[idx].strip()
        if text.startswith("## "):
            break
        if text.startswith("### "):
            current_section = text[4:].strip()
        elif text.startswith("- ") and current_section:
            bullet = text[2:].strip()
            if current_section == "This week":
                if ":" in bullet:
                    label, value = bullet.split(":", 1)
                    rows.append([f"This week - {label.strip()}", value.strip()])
                else:
                    rows.append([current_section, bullet])
            else:
                rows.append([current_section, bullet])
        idx += 1

    return rows


def extract_workflow(lines: list[str]) -> str:
    start_idx = find_heading(lines, "## Workflow") + 1
    numbered = []
    item_re = re.compile(r"^\d+\.\s+\*\*(.+?)\*\*\s*-\s*(.+)$")

    idx = start_idx
    while idx < len(lines):
        text = lines[idx].strip()
        if text.startswith("## "):
            break
        match = item_re.match(text)
        if match:
            numbered.append(f"{match.group(1)} - {match.group(2)}")
        idx += 1

    return " -> ".join(part.split(" - ", 1)[0] for part in numbered)


def style_sheet(sheet, freeze_cell: str = "A2") -> None:
    sheet.freeze_panes = freeze_cell
    for col_idx, _ in enumerate(sheet[1], start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in sheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = BODY_ALIGNMENT
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 48)


def write_table_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title)
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    style_sheet(sheet)


def write_overview_sheet(
    workbook: Workbook,
    workflow: str,
    board_rows: list[list[str]],
    milestone_rows: list[list[str]],
    queue_rows: list[list[str]],
) -> None:
    sheet = workbook.active
    sheet.title = "Overview"

    rows = [
        ["Project", "Aegis Protocol"],
        ["Mission", "Turn Aegis into a real-world, revenue-capable product."],
        ["Workflow", workflow],
        ["Markdown source", "docs/project-management/AEGIS_PROJECT_WORKBOOK.md"],
        ["Spreadsheet mirror", "docs/project-management/PROJECT_MANAGEMENT.xlsx"],
        ["Sync script", ".github/scripts/sync_project_management_xlsx.py"],
        ["Program epics", str(len(board_rows))],
        ["Milestones", str(len(milestone_rows))],
        ["First queue items", str(len(queue_rows))],
        ["Operating model", "docs/project-management/OPERATING_MODEL.md"],
    ]

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    notes_row = len(rows) + 2
    sheet.cell(notes_row, 1, "Operating rules")
    sheet.cell(notes_row, 1).fill = HEADER_FILL
    sheet.cell(notes_row, 1).font = HEADER_FONT
    sheet.cell(
        notes_row + 1,
        1,
        "Edit the markdown workbook first, then regenerate this spreadsheet to keep planning artifacts aligned.",
    )
    sheet.merge_cells(start_row=notes_row + 1, start_column=1, end_row=notes_row + 1, end_column=4)
    style_sheet(sheet, freeze_cell="A1")


def write_weekly_update_sheet(workbook: Workbook, weekly_rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet("Weekly Update")
    rows = [["Section", "Notes"]] + (
        weekly_rows
        if weekly_rows
        else [
            ["This week - completed", ""],
            ["This week - in progress", ""],
            ["This week - blocked", ""],
            ["Key decisions", ""],
            ["Risks", ""],
            ["Next week", ""],
        ]
    )

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    style_sheet(sheet)


def main() -> int:
    if not WORKBOOK_MD.exists():
        print(f"Missing workbook source: {WORKBOOK_MD}", file=sys.stderr)
        return 1

    workbook_lines = read_lines(WORKBOOK_MD)
    operating_model_lines = read_lines(OPERATING_MODEL_MD)

    board_headers, board_rows = parse_table_after_heading(workbook_lines, "## Program board")
    milestone_headers, milestone_rows = parse_table_after_heading(workbook_lines, "## Milestones")
    backlog_headers, backlog_rows = parse_backlog(workbook_lines)
    queue_headers, queue_rows = parse_numbered_list_after_heading(workbook_lines, "## First execution queue")
    commercial_headers, commercial_rows = parse_optional_table_after_heading_prefix(
        workbook_lines, "#### Commercial package"
    )
    target_headers, target_rows = parse_optional_table_after_heading_prefix(
        workbook_lines, "#### Ranked target-account list"
    )
    weekly_rows = parse_weekly_update(workbook_lines)
    workflow = extract_workflow(operating_model_lines)

    wb = Workbook()
    write_overview_sheet(wb, workflow, board_rows, milestone_rows, queue_rows)
    write_table_sheet(wb, "Program Board", board_headers, board_rows)
    write_table_sheet(wb, "Milestones", milestone_headers, milestone_rows)
    write_table_sheet(wb, "Backlog", backlog_headers, backlog_rows)
    write_table_sheet(wb, "Execution Queue", queue_headers, queue_rows)
    if commercial_headers and commercial_rows:
        write_table_sheet(wb, "Commercial Package", commercial_headers, commercial_rows)
    if target_headers and target_rows:
        write_table_sheet(wb, "Target Accounts", target_headers, target_rows)
    write_weekly_update_sheet(wb, weekly_rows)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
